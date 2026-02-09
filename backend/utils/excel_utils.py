"""
Excel工具类
"""
import os
import uuid
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook


def allowed_file(filename, allowed_extensions):
    """检查文件扩展名是否允许"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions


def generate_filename(original_filename):
    """生成唯一文件名"""
    ext = os.path.splitext(original_filename)[1]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    unique_id = uuid.uuid4().hex[:8]
    return f"{timestamp}_{unique_id}{ext}"


def read_excel(file_path, sheet_index=0, limit=None):
    """
    读取Excel文件
    
    Args:
        file_path: 文件路径
        sheet_index: 工作表索引，默认0
        limit: 限制返回行数，默认None返回全部
    
    Returns:
        list: 字典列表，每行数据
    """
    result = []
    
    try:
        # 使用pandas读取
        if limit:
            df = pd.read_excel(file_path, sheet_name=sheet_index, nrows=limit + 1)
        else:
            df = pd.read_excel(file_path, sheet_name=sheet_index)
        
        # 转换为字典列表
        result = df.to_dict('records')
        
        # 处理NaN值
        for row in result:
            for key, value in row.items():
                if pd.isna(value):
                    row[key] = None
                elif isinstance(value, (pd.Timestamp, datetime)):
                    row[key] = value.isoformat()
                elif isinstance(value, (int, float)) and not isinstance(value, bool):
                    # 处理数值
                    if isinstance(value, float) and value == value.long():
                        row[key] = int(value)
        
    except Exception as e:
        # 如果pandas读取失败，尝试使用openpyxl
        result = _read_excel_openpyxl(file_path, sheet_index, limit)
    
    return result


def _read_excel_openpyxl(file_path, sheet_index, limit):
    """使用openpyxl读取Excel（备选方案）"""
    result = []
    
    try:
        wb = load_workbook(file_path)
        sheet = wb.active if sheet_index == 0 else wb.worksheets[sheet_index]
        
        # 获取表头
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            headers.append(cell.value if cell.value else f"列{col}")
        
        # 读取数据
        row_count = 0
        for row in range(2, sheet.max_row + 1):
            if limit and row_count >= limit:
                break
            
            row_data = {}
            for col in range(1, len(headers) + 1):
                cell = sheet.cell(row=row, column=col)
                value = _get_cell_value(cell)
                row_data[headers[col - 1]] = value
            
            result.append(row_data)
            row_count += 1
        
    except Exception as e:
        raise Exception(f"读取Excel失败: {str(e)}")
    
    return result


def _get_cell_value(cell):
    """获取单元格值"""
    if cell is None:
        return None
    
    cell_type = cell.data_type
    if cell_type == 's':
        return cell.value
    elif cell_type == 'n':
        num_val = cell.value
        if num_val is not None and num_val == int(num_val):
            return int(num_val)
        return num_val
    elif cell_type == 'b':
        return cell.value
    elif cell_type == 'd':
        return cell.value.isoformat() if cell.value else None
    elif cell_type == 'f':
        return cell.value
    else:
        return cell.value


def generate_summary(file_path):
    """
    生成Excel文件摘要
    
    Args:
        file_path: 文件路径
    
    Returns:
        dict: 摘要信息
    """
    summary = {}
    
    try:
        # 使用pandas获取摘要
        df = pd.read_excel(file_path, nrows=1)
        
        summary['rowCount'] = 0
        summary['colCount'] = len(df.columns)
        summary['columns'] = list(df.columns)
        
        # 获取实际行数
        df_full = pd.read_excel(file_path)
        summary['rowCount'] = len(df_full)
        
    except Exception as e:
        raise Exception(f"生成摘要失败: {str(e)}")
    
    return summary


def write_excel(file_path, headers, data):
    """
    写入Excel文件
    
    Args:
        file_path: 文件路径
        headers: 表头列表
        data: 数据列表（字典）
    """
    import openpyxl
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header)
            if value is not None:
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 保存文件
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb.save(file_path)

